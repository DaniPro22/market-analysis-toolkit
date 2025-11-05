import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def exportar_a_excel(ruta_salida: str, hojas: dict, imagenes: list = None):
    """
    Exporta varias hojas a Excel, con control de tamaño y gráficos automáticos.
    Si se pasan imágenes en 'imagenes', se insertan al final del archivo Excel.
    """

    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)

    # 1️⃣ Exportar todas las hojas
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        for nombre, contenido in hojas.items():
            if isinstance(contenido, pd.DataFrame):
                filas, columnas = contenido.shape
                if filas > 1_048_000:
                    print(f"⚠️ Hoja '{nombre}' demasiado grande ({filas:,} filas). Guardando parcial en Excel y completa en CSV.")
                    contenido.head(50_000).to_excel(writer, sheet_name=nombre[:31], index=False)
                    csv_path = ruta_salida.replace(".xlsx", f"_{nombre}.csv")
                    contenido.to_csv(csv_path, index=False, encoding="utf-8-sig")
                    print(f"📁 CSV completo guardado en: {csv_path}")
                else:
                    contenido.to_excel(writer, sheet_name=nombre[:31], index=False)
            else:
                # Contenido tipo texto o Markdown
                df_texto = pd.DataFrame({"Contenido": [str(contenido)]})
                df_texto.to_excel(writer, sheet_name=nombre[:31], index=False)

    # 2️⃣ Reabrir para insertar gráficos
    wb = load_workbook(ruta_salida)

    # a) Gráficos automáticos (series temporales)
    for nombre, contenido in hojas.items():
        if not isinstance(contenido, pd.DataFrame) or contenido.empty:
            continue

        posibles_tiempo = ["AÑO", "YEAR", "FECHA", "DATE", "TIME_PERIOD"]
        col_tiempo = next((c for c in posibles_tiempo if c in contenido.columns), None)
        col_valor = "VALOR" if "VALOR" in contenido.columns else None

        if col_tiempo and col_valor:
            plt.figure(figsize=(6, 3))
            plt.plot(contenido[col_tiempo], contenido[col_valor], marker="o", linewidth=1.2)
            plt.title(f"{nombre} - Evolución temporal")
            plt.xlabel(col_tiempo)
            plt.ylabel(col_valor)
            plt.grid(True, linestyle="--", alpha=0.6)
            plt.tight_layout()

            img_path = f"{ruta_salida.replace('.xlsx', '')}_{nombre}.png"
            plt.savefig(img_path, dpi=120)
            plt.close()

            if nombre[:31] in wb.sheetnames:
                ws = wb[nombre[:31]]
                img = Image(img_path)
                img.width = 480
                img.height = 240
                ws.add_image(img, "H2")

    # b) Imágenes personalizadas (ej. simulaciones de cartera)
    if imagenes:
        for img_path in imagenes:
            if not os.path.exists(img_path):
                continue
            nombre_hoja = os.path.splitext(os.path.basename(img_path))[0][:31]
            if nombre_hoja not in wb.sheetnames:
                # Crear hoja para la imagen si no existe
                ws = wb.create_sheet(nombre_hoja)
            else:
                ws = wb[nombre_hoja]
            img = Image(img_path)
            img.width = 600
            img.height = 320
            ws.add_image(img, "A2")

    wb.save(ruta_salida)
    wb.close()

    print(f"✅ Archivo Excel exportado con gráficos: {ruta_salida}")

